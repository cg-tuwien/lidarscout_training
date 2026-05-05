import os
import openpyxl as pyx
import numpy as np
import math
import typing
import imageio.v3 as iio
from PIL import Image, ImageDraw, ImageFont

in_path = r'./results/'

datasets = [
    'ca_13', 
    # 'swisssurface3d', 'Bund_BoraPk', 'ID15_Bunds', 'NZ23_Gisborne_subsets_BF44',
    # 'NZ23_Gisborne_subsets_BG41_0to23', 'NZ23_Gisborne_subsets_BG41_24to50',
    # 'notoeast1', 'sitn',
            ]

runs = [
    'ipes_cnn',
    # 'ipes_gan',
    # 'gan',
    # 'ipes_cnn', 'ipes_cnn_colorizer', 'ipes_cnn_only_nn', 'ipes_cnn_only_lin', 'ipes_dctnet',
    # 'ipes_cnn_allstar', 'ipes_unet', 'ipes_rast',
    # 'ipes_interp_cubic', 'ipes_interp_linear', 'ipes_interp_rast_hqsplat_mean',
        ]

METRIC_SPECS = [
    ('hm_rmse_ms_mean', 'hm_rmse_ms', False),
    ('hm_gradient_rmse_mean', 'hm_gradient_rmse', False),
    # ('hm_lpips_mean', 'hm_lpips', False), 
    ('rgb_psnr_mean', 'rgb_psnr', True),
    ('rgb_lpips_mean', 'rgb_lpips', False),
    # ('rgb_ssim_mean', 'rgb_ssim', True),
    # ('rgb_flip_mean', 'rgb_flip', False),
    ('rgb_gradient_rmse_mean', 'rgb_gradient_rmse', False),
]

IMAGE_METRIC_SPECS = [
    ('rgb_psnr_mean', 'rgb_psnr', 'rgb_psnr', True),
    ('hm_rmse_ms_mean', 'hm_rmse', 'hm_rmse_ms', False),
    ('rgb_gradient_rmse_mean', 'rgb_grad_rmse', 'rgb_gradient_rmse', False),
    ('hm_gradient_rmse_mean', 'hm_grad_rmse', 'hm_gradient_rmse', False),
    # ('hm_lpips_mean', 'hm_lpips', 'hm_lpips', False),
    ('rgb_lpips_mean', 'rgb_lpips', 'rgb_lpips', False),
]

header = [
    'file',
    *[metric_name for _, metric_name, _ in METRIC_SPECS],
]

join_str = '\t '
value_format_str = '{:.4f}'

# header
header_joined = join_str.join(header)
print(header_joined)

# image stitching config
image_rel_dir = os.path.join('laz_minimal', 'test', '00_hm_rgb_fig_input')
image_file_names = [
    'ca_13_0_10_b0_00_hm_rgb_fig_input.png',
    'ca_13_10_20_b0_00_hm_rgb_fig_input.png',
    'ca_13_20_30_b0_00_hm_rgb_fig_input.png',
]


def _to_rgb(img: np.ndarray) -> np.ndarray:
    if img.ndim == 2:
        return np.stack([img, img, img], axis=-1)
    if img.ndim == 3 and img.shape[2] == 1:
        return np.repeat(img, 3, axis=2)
    if img.ndim == 3 and img.shape[2] >= 3:
        return img[:, :, :3]
    raise ValueError(f'Unexpected image shape: {img.shape}')


def _crop_patch_figure(img: np.ndarray) -> np.ndarray:
    h, w = img.shape[:2]
    target_w = min(w, h)
    target_h = max(1, h // 2)

    x0 = max(0, w - target_w)
    y0 = max(0, h - target_h)

    cropped = img[y0:y0 + target_h, x0:x0 + target_w]
    return cropped


def _add_label(img: np.ndarray, lines: typing.List[str], pad: int = 6) -> np.ndarray:
    pil_img = Image.fromarray(img)
    draw = ImageDraw.Draw(pil_img)
    font = ImageFont.load_default()

    line_heights = []
    max_line_w = 0
    for line in lines:
        bbox = draw.textbbox((0, 0), line, font=font)
        line_w = bbox[2] - bbox[0]
        line_h = bbox[3] - bbox[1]
        max_line_w = max(max_line_w, line_w)
        line_heights.append(line_h)

    text_h = int(sum(line_heights) + max(0, len(lines) - 1) * 2)
    header_h = text_h + 2 * pad
    out_w = max(pil_img.width, max_line_w + 2 * pad)

    out = Image.new('RGB', (int(out_w), int(pil_img.height + header_h)), color=(255, 255, 255))
    out.paste(pil_img, (0, header_h))

    draw_out = ImageDraw.Draw(out)
    y = pad
    for line, line_h in zip(lines, line_heights):
        draw_out.text((pad, y), line, fill=(0, 0, 0), font=font)
        y += line_h + 2

    return np.asarray(out)


def _metric_bg_color(value: typing.Optional[float], vmin: float, vmax: float, higher_is_better: bool) -> typing.Tuple[int, int, int]:
    if value is None:
        return 255, 255, 255

    if vmax <= vmin:
        score = 0.5
    else:
        norm = (value - vmin) / (vmax - vmin)
        norm = max(0.0, min(1.0, norm))
        score = norm if higher_is_better else 1.0 - norm

    red = int(round(255.0 * (1.0 - score)))
    green = int(round(255.0 * score))
    blue = 140
    return red, green, blue


def _load_monospace_font(size: int = 14):
    font_candidates = [
        r'C:\Windows\Fonts\consola.ttf',
        r'C:\Windows\Fonts\consolab.ttf',
        'DejaVuSansMono.ttf',
    ]
    for font_path in font_candidates:
        try:
            return ImageFont.truetype(font_path, size=size)
        except Exception:
            pass
    return ImageFont.load_default()


def _fmt_fixed(value: typing.Optional[float], width: int = 8, precision: int = 4) -> str:
    if value is None:
        return 'None'.rjust(width)
    return f'{value:>{width}.{precision}f}'


def _ellipsize_left(text: str, max_len: int) -> str:
    if max_len <= 0:
        return ''
    if len(text) <= max_len:
        return text
    if max_len <= 3:
        return text[-max_len:]
    return '...' + text[-(max_len - 3):]


def _add_label_colored_metrics(
        img: np.ndarray,
        run_name: str,
        means: typing.Dict[str, typing.Optional[float]],
        metric_ranges: typing.Dict[str, typing.Tuple[float, float]],
        run_name_width: int,
        value_fmt: str,
    pad: int = 2) -> np.ndarray:
    pil_img = Image.fromarray(img)

    metric_values = {display_name: means.get(mean_name) for mean_name, display_name, _, _ in IMAGE_METRIC_SPECS}
    metric_texts = {
        display_name: _fmt_fixed(metric_values[display_name], width=max(len('40.0000'), len('None')), precision=4)
        for _, display_name, _, _ in IMAGE_METRIC_SPECS
    }
    out_w = pil_img.width
    selected_font = _load_monospace_font(size=14)
    draw_tmp = ImageDraw.Draw(Image.new('RGB', (1, 1), color=(255, 255, 255)))

    def _text_w(text: str, font: typing.Any) -> int:
        bbox = draw_tmp.textbbox((0, 0), text, font=font)
        return int(bbox[2] - bbox[0])

    def _chunk_text(text: str, max_chars: int) -> typing.List[str]:
        if max_chars <= 1:
            return [text]
        chunks = []
        start = 0
        while start < len(text):
            chunks.append(text[start:start + max_chars])
            start += max_chars
        return chunks if len(chunks) > 0 else ['']

    metric_name_w = max(len(display_name) for _, display_name, _, _ in IMAGE_METRIC_SPECS)
    value_w = max(len('40.0000'), len('None'))
    block_name = f"{'x' * metric_name_w}="
    block_value = 'x' * value_w
    sep = ' | '

    chosen_font = selected_font
    chosen_line_h = 0
    chosen_run_lines: typing.List[str] = []
    chosen_cols = 1
    chosen_block_name_w = 0
    chosen_sep_w = 0

    for font_size in [14, 13, 12, 11, 10, 9, 8]:
        font = _load_monospace_font(size=font_size)
        sample_bbox = draw_tmp.textbbox((0, 0), 'Ag', font=font)
        line_h = int(sample_bbox[3] - sample_bbox[1])

        block_name_w_px = _text_w(block_name, font)
        block_w_px = _text_w(block_name + block_value, font)
        sep_w_px = _text_w(sep, font)
        available_w = max(1, out_w - 2 * pad)

        cols = 1
        while True:
            next_cols = cols + 1
            next_required = next_cols * block_w_px + (next_cols - 1) * sep_w_px
            if next_required <= available_w:
                cols = next_cols
            else:
                break

        # Prefer fewer metric rows for readability while keeping run-name wrapping.
        char_w_px = max(1, _text_w('M', font))
        max_chars = max(1, available_w // char_w_px)
        run_lines = _chunk_text(run_name, max_chars)

        chosen_font = font
        chosen_line_h = line_h
        chosen_run_lines = run_lines
        chosen_cols = cols
        chosen_block_name_w = block_name_w_px
        chosen_sep_w = sep_w_px
        break

    metric_rows = int(math.ceil(len(METRIC_SPECS) / chosen_cols))
    total_text_lines = len(chosen_run_lines) + metric_rows
    header_h = total_text_lines * chosen_line_h + (total_text_lines + 1) * pad

    out = Image.new('RGB', (out_w, pil_img.height + header_h), color=(255, 255, 255))
    out.paste(pil_img, (0, header_h))
    draw_out = ImageDraw.Draw(out)

    y = pad
    for run_line in chosen_run_lines:
        draw_out.text((pad, y), run_line, fill=(0, 0, 0), font=chosen_font)
        y += chosen_line_h + pad

    block_w_px = _text_w(block_name + block_value, chosen_font)
    col_stride = block_w_px + chosen_sep_w
    for row in range(metric_rows):
        x = pad
        for col in range(chosen_cols):
            metric_idx = row * chosen_cols + col
            if metric_idx >= len(IMAGE_METRIC_SPECS):
                break

            mean_name, display_name, source_metric_name, higher_is_better = IMAGE_METRIC_SPECS[metric_idx]
            name_text = f'{display_name:<{metric_name_w}}='
            value_text = metric_texts[display_name]
            bg_color = _metric_bg_color(metric_values[display_name], *metric_ranges[mean_name], higher_is_better=higher_is_better)

            draw_out.text((x, y), name_text, fill=(0, 0, 0), font=chosen_font)
            val_x = x + chosen_block_name_w
            draw_out.rectangle([val_x, y, val_x + _text_w(value_text, chosen_font), y + chosen_line_h], fill=bg_color)
            draw_out.text((val_x, y), value_text, fill=(0, 0, 0), font=chosen_font)
            x += col_stride

        y += chosen_line_h + pad

    return np.asarray(out)


def _make_grid(images: typing.List[np.ndarray], pad: int = 10, bg_value: int = 255) -> np.ndarray:
    if len(images) == 0:
        raise ValueError('No images to stitch')

    max_h = max(img.shape[0] for img in images)
    max_w = max(img.shape[1] for img in images)
    n = len(images)
    cols = int(math.ceil(math.sqrt(n)))
    rows = int(math.ceil(n / cols))

    grid_h = rows * max_h + (rows + 1) * pad
    grid_w = cols * max_w + (cols + 1) * pad
    grid = np.full((grid_h, grid_w, 3), bg_value, dtype=np.uint8)

    for idx, img in enumerate(images):
        row = idx // cols
        col = idx % cols
        y0 = pad + row * (max_h + pad)
        x0 = pad + col * (max_w + pad)

        y_off = (max_h - img.shape[0]) // 2
        x_off = 0

        ys = y0 + y_off
        xs = x0 + x_off
        grid[ys:ys + img.shape[0], xs:xs + img.shape[1], :] = img

    return grid


def _get_metric_means(excel_file: str) -> typing.Dict[str, typing.Optional[float]]:
    wb = pyx.load_workbook(excel_file)
    ws = wb.active
    if ws is None:
        wb.close()
        return {mean_name: None for mean_name, _, _ in METRIC_SPECS}

    header_row = [cell.value for cell in ws[1]]
    header_to_col = {name: idx + 1 for idx, name in enumerate(header_row) if name is not None}

    average_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == 'AVERAGE':
            average_row = row_idx
            break
    if average_row is None:
        average_row = ws.max_row + 1

    means: typing.Dict[str, typing.Optional[float]] = {}
    for mean_name, metric_name, _ in METRIC_SPECS:
        col_id = header_to_col.get(metric_name)
        if col_id is None:
            means[mean_name] = None
            continue

        cell_vals_raw = [ws.cell(row=row_idx, column=col_id).value for row_idx in range(2, average_row)]
        cell_vals = [float(v) for v in cell_vals_raw if isinstance(v, (int, float, np.floating))]

        # fix broken PSNR
        if metric_name == 'rgb_psnr':
            cell_vals = [v - 48.131 if v > 30 else v for v in cell_vals]

        if len(cell_vals) == 0:
            means[mean_name] = None
            continue

        cell_vals_np = np.asarray(cell_vals, dtype=np.float64)
        means[mean_name] = float(np.mean(cell_vals_np))

    wb.close()
    return means


metrics_by_run: typing.Dict[str, typing.Dict[str, typing.Optional[float]]] = {}
metrics_by_run_dataset: typing.Dict[typing.Tuple[str, str], typing.Dict[str, typing.Optional[float]]] = {}

for dataset in datasets:
    # get all excel files recursively
    excel_files = []
    for root, dirs, files in os.walk(in_path):
        for file in files:
            if file.endswith('.xlsx') and file.find(dataset) != -1:
                contains_run = any([file.find(run) != -1 for run in runs])
                if contains_run:
                    excel_files.append(os.path.join(root, file))

    # read all excel files
    parsed_excel_files = []
    for excel_file in excel_files:
        try:
            means = _get_metric_means(excel_file)
        except Exception:
            print(f'Error loading {excel_file}')
            continue

        vals_to_print = []
        for mean_name, _, _ in METRIC_SPECS:
            value = means.get(mean_name)
            vals_to_print.append(value_format_str.format(value) if value is not None else 'None')

        file_name = os.path.basename(excel_file)
        if file_name.startswith('metrics_'):
            file_name = file_name[len('metrics_'):]
        run_name = file_name.split('_test_')[0]
        parsed_excel_files.append((run_name, excel_file, means, vals_to_print))
        metrics_by_run[run_name] = means
        metrics_by_run_dataset[(run_name, dataset)] = means

    for run_name, excel_file, means, vals_to_print in sorted(parsed_excel_files, key=lambda item: item[0]):
        vals_joined = join_str.join(vals_to_print)
        print(f'{os.path.basename(excel_file)[13:]} {join_str} {vals_joined}', flush=True)

# build labeled patch comparison images per dataset
run_dirs = sorted([d for d in os.listdir(in_path) if os.path.isdir(os.path.join(in_path, d))])
for dataset in datasets:
    dataset_image_files = [f for f in image_file_names if f.startswith(dataset + '_')]
    if len(dataset_image_files) == 0:
        print(f'No image file pattern configured for dataset {dataset}')
        continue

    for image_file_name in dataset_image_files:
        run_entries = []
        for run_dir in run_dirs:
            if not any(run in run_dir for run in runs):
                continue

            img_file = os.path.join(in_path, run_dir, image_rel_dir, image_file_name)
            if not os.path.isfile(img_file):
                continue

            try:
                img = iio.imread(img_file)
                img = _to_rgb(np.asarray(img))
                img = _crop_patch_figure(img)
                if img.dtype != np.uint8:
                    img = np.clip(img, 0, 255).astype(np.uint8)

                means = metrics_by_run_dataset.get((run_dir, dataset), metrics_by_run.get(run_dir, {}))
                if means is None:
                    means = {}
                run_entries.append((run_dir, img, means))
            except Exception as exc:
                print(f'Error loading image {img_file}: {exc}')

        images_labeled = []
        if len(run_entries) > 0:
            run_entries.sort(key=lambda entry: entry[0])
            metric_ranges = {}
            run_name_width = max(len(entry[0]) for entry in run_entries)
            for mean_name, _, source_metric_name, _ in IMAGE_METRIC_SPECS:
                vals = [entry[2].get(source_metric_name) for entry in run_entries]
                vals = [v for v in vals if v is not None]
                if len(vals) == 0:
                    metric_ranges[mean_name] = (0.0, 1.0)
                else:
                    metric_ranges[mean_name] = (float(min(vals)), float(max(vals)))

            for run_dir, img, means in run_entries:
                labeled = _add_label_colored_metrics(
                    img=img,
                    run_name=run_dir,
                    means=means,
                    metric_ranges=metric_ranges,
                    run_name_width=run_name_width,
                    value_fmt=value_format_str,
                )
                images_labeled.append(labeled)

        if len(images_labeled) > 0:
            stitched = _make_grid(images_labeled)
            image_file_stem, _ = os.path.splitext(image_file_name)
            image_out_file = os.path.join(in_path, f'patch_comp_{image_file_stem}.png')
            iio.imwrite(image_out_file, stitched)
            print(f'Wrote {image_out_file} ({len(images_labeled)} images)')
        else:
            print(f'No patch images found for dataset={dataset}, file={image_file_name}.')