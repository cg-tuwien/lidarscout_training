import math

import numpy as np
import os
import typing

from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import pandas as pd


def make_test_report(shape_names: list, results: typing.Union[list, dict],
                     output_file: str, output_names: list, is_dict=True,
                     metrics_keys_to_log=frozenset(['abs_dist_rms', 'accuracy', 'precision', 'recall', 'f1_score']),
                     low_metrics_better: typing.Sequence[bool] = frozenset([True, False, False, False, False])):
    import pandas as pd
    from torch import stack

    headers = ['Shape', 'Loss total'] + output_names + list(metrics_keys_to_log)
    all_low_metrics_better = [True] * (1 + len(output_names)) + list(low_metrics_better)

    if not is_dict:
        loss_total = [r[0] for r in results]
        loss_components = [r[1] for r in results]
        metrics_dicts = [r[2] for r in results]
        metrics_lists = []
        for m in metrics_keys_to_log:
            metrics_list = [md[m] for md in metrics_dicts]
            metrics_lists.append(metrics_list)
        metrics = np.array(metrics_lists).transpose()
    else:
        loss_total = results['loss'].detach().cpu()
        loss_components = results['loss_components_mean'].detach().cpu()
        if len(loss_components.shape) == 1:
            loss_components = loss_components.unsqueeze(1)
        metrics = stack([results[k] for k in metrics_keys_to_log]).transpose(0, 1).detach().cpu()

        if len(loss_total.shape) == 2:  # DP -> squeeze
            loss_total = loss_total.squeeze(-1)
            metrics = metrics.squeeze(-1)

    data = [[results['pc_file_in'][i]] + [loss_total[i].item()] + loss_components[i].tolist() + metrics[i].tolist()
            for i in range(len(loss_total))]
    df = pd.DataFrame(data=data, columns=headers)
    df = df.set_index('Shape')

    export_xlsx(df=df, low_metrics_better=all_low_metrics_better, output_file=output_file,
                add_stats=True, header=True, independent_cols=True)

    loss_total_mean = np.mean(np.array(loss_total))
    return loss_total_mean, np.nanmean(metrics, axis=0)


def export_xlsx(df: 'pd.DataFrame', low_metrics_better: typing.Union[None, typing.Sequence[bool], bool],
                output_file: str, add_stats=True, header=True, independent_cols=True):
    import datetime
    from source.base import fs

    # export with conditional formatting and average
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.utils.cell import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule
    wb = Workbook()
    ws: Workbook = wb.active

    df_export = df.copy()
    df_export.reset_index(inplace=True)  # revert index to normal column to get rid of extra header row
    for r in dataframe_to_rows(df_export, index=False, header=header):
        ws.append(r)

    # no direction given, assume near 0 or near 1 results
    if low_metrics_better is None:
        cols = df.to_numpy()
        cols = np.vectorize(lambda x: x.timestamp() if isinstance(x, datetime.datetime) else x)(cols)
        cols_mean = np.nanmean(cols, axis=0)
        if not independent_cols:
            cols_mean = np.mean(cols_mean)  # scalar for dependent cols
        low_metrics_better = np.logical_or(cols_mean > 1.0, cols_mean < 0.5)

    top_row = 2
    col_ids = df.index.shape[1] if len(df.index.shape) > 1 else 1
    ws.freeze_panes = '{}{}'.format(get_column_letter(col_ids + 1), top_row)
    bottom_row = df.shape[0] + top_row - 1
    if add_stats:
        for di in range(df.shape[1]):
            # without this, openpyxl will turn unknown functions like "IF" into "@IF" and Excel will not recognize it
            xl_fn = '_xlfn.'

            column = col_ids + 1 + di
            column_letter = get_column_letter(column)
            range_str = '{}{}:{}{}'.format(column_letter, top_row, column_letter, bottom_row)

            # doesn't work for some reason, produces a "@" in front of the range,
            # which fails if the cell is in the same column
            ignore_zeros_str = '{}IF({}<>0, {})'.format(xl_fn, range_str, range_str)

            ws.cell(row=bottom_row + 1, column=column).value = '=AVERAGE({})'.format(range_str)
            ws.cell(row=bottom_row + 2, column=column).value = '=MEDIAN({})'.format(range_str)
            ws.cell(row=bottom_row + 3, column=column).value = '={}STDEV.P({})'.format(xl_fn, range_str)

        # Stat names
        ws.cell(row=bottom_row + 1, column=1).value = 'AVERAGE'
        ws.cell(row=bottom_row + 2, column=1).value = 'MEDIAN'
        ws.cell(row=bottom_row + 3, column=1).value = 'STDEV'

    def add_formatting_rule(col_start_id, row_start_id, col_end_id, row_end_id, lower_is_better):
        col_start_letter = get_column_letter(col_start_id)
        col_end_letter = get_column_letter(col_end_id)
        col_range_str = '{col_start}{row_start}:{col_end}{row_end}'.format(
            col_start=col_start_letter, row_start=row_start_id, col_end=col_end_letter, row_end=row_end_id)
        if lower_is_better:  # error here means that this list has an invalid length
            start_color = 'FF00AA00'
            end_color = 'FFAA0000'
        else:
            end_color = 'FF00AA00'
            start_color = 'FFAA0000'
        rule = ColorScaleRule(start_type='percentile', start_value=0, start_color=start_color,
                              mid_type='percentile', mid_value=50, mid_color='FFFFFFFF',
                              end_type='percentile', end_value=100, end_color=end_color)
        ws.conditional_formatting.add(col_range_str, rule)

        # highlight optimum
        from openpyxl.formatting.rule import FormulaRule
        from openpyxl.styles import Font

        asc_desc = 'MIN' if lower_is_better else 'MAX'
        # should be like =H2=MIN(H$2:H$11)
        formula = '={col_start}{row_start}={func}({col_start}${row_start}:{col_end}${row_end})'.format(
            col_start=col_start_letter, row_start=row_start_id, func=asc_desc,
            col_end=col_end_letter, row_end=row_end_id)
        rule = FormulaRule(formula=(formula,), font=Font(underline='single'))
        ws.conditional_formatting.add(col_range_str, rule)

    # color scale over shapes
    if independent_cols:
        bottom_row_formatting = bottom_row + (2 if add_stats else 0)  # not for STDEV
        for col in range(df.shape[1]):
            if not np.isnan(low_metrics_better[col]):
                add_formatting_rule(col_start_id=col+col_ids+1, row_start_id=top_row,
                                    col_end_id=col+col_ids+1, row_end_id=bottom_row_formatting,
                                    lower_is_better=low_metrics_better[col])
    else:  # dependent cols
        for shape_id in range(df.shape[0]):
            row = top_row + shape_id
            add_formatting_rule(col_start_id=col_ids+1, row_start_id=row,
                                col_end_id=df.shape[1]+col_ids+1, row_end_id=row,
                                lower_is_better=low_metrics_better)

        # color scale over stats (horizontal)
        lower_better = [low_metrics_better] * 2 + [True]  # lower stdev is always better, mean and avg depend on metric
        for stat_id in range(3):
            row = bottom_row + 1 + stat_id
            add_formatting_rule(col_start_id=col_ids+1, row_start_id=row,
                                col_end_id=df.shape[1] + col_ids+1, row_end_id=row,
                                lower_is_better=lower_better[stat_id])

    fs.make_dir_for_file(output_file)
    wb.save(output_file)
